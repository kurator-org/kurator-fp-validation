# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

__author__ = "Robert A. Morris"
__copyright__ = "Copyright 2016 President and Fellows of Harvard College"
__version__ = "OutcomeStats.py 2017-02-20T18:52:21-0500"

from actor_decorator import python_actor
from OutcomeFormats import *

import json
import sys
#import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import OpenpyxlStyle
from OpenpyxlStyle import style_range

#import openpyxl
import configparser
import Args

class OutcomeStats:
#   def __init__(self, workbook, worksheet,infile, outfile, configFile, origin1, origin2):
   def __init__(self, configfile):

      #with open(args.getInfile()) as data_file:
      #           fpAkkOutput=json.load(data_file)

      #with open(args.getInfile()) as data_file:
      #   self.fpa=json.load(data_file)

      config = configparser.ConfigParser()
      config.sections()
#      self.configFile =configFile
      self.configFile = configfile
#      self.configFile='stats.ini'
      config.read(self.configFile)
     #### self.validators =eval( config['DEFAULT']['validators'])
      validators = ("ScientificNameValidator","DateValidator",  "GeoRefValidator","BasisOfRecordValidator") #row order in output
      outcomes = ("CORRECT","CURATED","FILLED_IN", "UNABLE_DETERMINE_VALIDITY",  "UNABLE_CURATE") #col order in output
      self.maxlength= max(len(s) for s in validators)

     #### self.outcomes = eval(config['DEFAULT']['outcomes'])
      self.max1= max(len(s) for s in validators)
      self.max2= max(len(t) for t in outcomes)
      self.maxlength = max(self.max1,self.max2)
      #self.fpa = {}
      #infile = 'occurrence_qc.json' #for now

      #self.numRecords = len(self.fpa)

   def getOutcomes(self) :
      return self.outcomes
   def getValidators(self) :
      return self.validators
   def getMaxLength(self):
      return self.maxlength

   def initStats(self,outcomes) :
      stats = {}
      for outcome in outcomes:
          stats[outcome] = 0
      return stats
   
   def initValidatorStats(self,validators, outcomes) :
      stats = {}
      for v in validators :
         stats[v] = self.initStats(outcomes)
      return stats
   
   def updateValidatorStats(self,fpa, stats, record)  :
      data=fpa[record]["Markers"]
   #   print("in updateValidatorStats[",record,"]")
      for data_k, data_v in data.items() :
         for stats_k, stats_v in stats.items() :
            if (stats_k == data_k):
               stats[stats_k][data_v] += 1
      return stats
   
   #typed parameter requires python3
   def createStats(self, fpa, normalize):
      validatorStats = self.initValidatorStats(self.validators, self.outcomes)
      for record in range(len(fpa)):
         self.updateValidatorStats(fpa, validatorStats, record) 
      if normalize == True :
         self.normalizeStats(fpa,validatorStats)
      return validatorStats
   
   def normalizeStats(self,fpa,stats):
      #fpa is dict loaded from FP-Akka json output
      #divide outcome counts by occurrence counts
      count=len(fpa)
      count_f= float(count)
   #   if (count <= 0) return(-1)
      for validator,outcomes in stats.items():
         stat=stats[validator]
         for k,v in stat.items():
            v = v/count_f
            stat[k] = format(v, '.4f')
   #         print("yy:",stats[validator])
   #   print("in normalize stats=",stats)
      return stats
   
   def stats2CSV(self, stats, outfile, outcomes, validators):
      import csv
      with open(outfile, 'w') as csvfile:
         o=list(outcomes)
         o.insert(0,"Validator")
         fieldnames=tuple(o)
         writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
         writer.writeheader()
         for v in validators:
            row = stats[v]
            row['Validator'] = v
            writer.writerow(row)
   
#   def initWorkbook(outfile):
   def initWorkbook(options):
      """
      Returns a workbook to be written to **outfile**
      """
#      workbook = xlsxwriter.Workbook(outfile)
      outfile = optdict['outputfile']
      workbook = openpyxl.Workbook(outfile)
      return workbook
   
   #doesn't belong in this class
   
#   def stats2XLSX(self, workbook, worksheet, formats, stats, origin, outcomes, validators):
   def stats2XLSX(self, optdict):

      # David: construct origin from dictionary values for origincolumn and originrow
      origin = [optdict['origincolumn'], optdict['originrow']]

   #   print("fmts=",formats)
#      bold = workbook.add_format({'bold': True})
      
   #   print("stats=",stats)
   #   print("outcomes=", outcomes)
   #   print('origin=',origin)
      headRow = origin[0]


      headCol = origin[1]
####  # set cell_range for call to OpenpyxlStyle.style_range(....)
#      ws=worksheet
      wb = optdict['workbook']
      ws = optdict['worksheet']
      border = optdict['border']
      fill = optdict['fill']
      al = optdict['alignment']

###      worksheet.write(headRow ,headCol,"Validator",bold)
###      font = Font(name='Calibri',size=11,bold='True')
      font = optdict['font']
      cell_range = ws.cell(column=headCol, row=headRow, value="Validator")
      print("HELLO FROM OutcomeStats L163\n")
#      OpenpyxlStyle.style_range(ws, cell_range, border=border, fill=fill, font=font, alignment=al)
      OpenpyxlStyle.style_range(optdict)
####
      validators = ("ScientificNameValidator","DateValidator",  "GeoRefValidator","BasisOfRecordValidator") #row order in output
      outcomes = ("CORRECT","CURATED","FILLED_IN", "UNABLE_DETERMINE_VALIDITY",  "UNABLE_CURATE") #col order in output
      
#      wb = Workbook()
      ws = wb.active
      my_cell = ws['B6']
      my_cell.value = "Validator"
#      OpenpyxlStyle.style_range(ws,'B2:F4', border = border, fill=fill, font=font, alignment=al)
      OpenpyxlStyle.style_range(optdict)
     # wb.save(optdict['outputfile'])
#      sys.exit()
#      return
      
      print("HELLO FROM OutcomeStats L180\n")
      formatdict={}
      outcomeFormats = OutcomeFormats(formatdict)
      xxx = outcomeFormats.initFormats(formatdict) #null for now
      print("HELLO from L 188 outcomeFormats =", outcomeFormats)
      sys.exit()

      return
      for str in outcomes :
         col=1+headCol+outcomes.index(str) #insure order is as in outcomes list
###         worksheet.write(headRow,col, str, bold) #write col header
      print("HELLO FROM OutcomeStats L184\n")
      for k, v in stats.items():
         col = headCol;
 ###        print("at L187 key=",k,"val=", v, "thecol=",col)
         row = 1+headRow+validators.index(k) #put rows in order of the validators list
         print('OutcomeStats at L179 row=',row, 'thecol=',col, 'k=',k)
   #      print("row=",row)
###         worksheet.write(row,0,k) #write validator name
         worksheet.write(row,col,k) #write validator name
         #write data for each validator in its own row
         for outcome, statval in v.items():
            col = headCol + 1 + outcomes.index(outcome) #put cols in order of the outcomes list
            worksheet.write(row, col, statval,formats.get(outcome))


@python_actor
###def outcomestats(inputfile, outputfile, configfile, origincolumn, originrow):
#def outcomestats(inputfile, outputfile, configfile, originrow, origincolumn):
def outcomestatsOnData(optdict):
   # load entire jason file. (Note: syntactically it is a Dictionary !!! )
 ###  print('\nL210 in outcomestatsOnData. optdict=', optdict)
   inputfile = optdict['inputfile']  #jason

   # David: assigning values from optdict to local variables that used to be function args
   configfile = optdict['configfile']
   origincolumn = optdict['origincolumn']
   originrow = optdict['originrow']

   with open(inputfile) as data_file:
      fpAkkaOutput = json.load(data_file)
#   origin1 = [origincolumn,originrow]

#   wb = Workbook()  # xlsxwriteropenpyxl model of an xlsx spreadsheet
#   worksheet = workbook.add_worksheet()  # should supply worksheet name, else defaults
#   ws = wb.create_sheet("stats",0)  ##should be different openpyxl worksheet for each workflow?

   #   stats = OutcomeStats(workbook,worksheet,data_file,outfile,configFile,origin1,origin2)
   ###stats = OutcomeStats(configfile)
   ###    worksheet.set_column(0, len(stats.getOutcomes()), 3 + stats.getMaxLength())
   ###    worksheet.set_column(origincolumn, len(stats.getOutcomes()), 3 + stats.getMaxLength())
#   worksheet.set_column(origincolumn, len(stats.getOutcomes()), 3 + stats.getMaxLength())
  # ws.set_column(origincolumn, len(stats.getOutcomes()), 3 + stats.getMaxLength())
###   d = ws.cell(row=4, column=2, value=10)
#   ws.cell(column=origincolumn, row=originrow, value='hello')
###   print('L199=')
   #   print(stats.getOutcomes())
   outcomeFormats = OutcomeFormats({})
#xw   formats = outcomeFormats.initFormats(workbook)  # shouldn't be attr of main class
#   formats = outcomeFormats.initFormats(wb,ws)  # shouldn't be attr of main class
### formats map outcome to openpyxl style data
   formatsDict = optdict #for now
   formats = outcomeFormats.initFormats(formatsDict)  #
   ###################################################
   #####createStats and stats2XLSX comprise the main #
   # processor filling the spreadheet cells       ####
   ###################################################
   # if stats are normalized, results are divided by number of records
   # otherwise, cells show total of the number of each outcome in the appropriate column
   normalized = True
###   validatorStats = stats.createStats(fpAkkaOutput, ~normalized)
###   validatorStatsNormalized = stats.createStats(fpAkkaOutput, normalized)

####   outcomes = stats.getOutcomes()
   validators = ("ScientificNameValidator","DateValidator",  "GeoRefValidator","BasisOfRecordValidator") #row order in output
   outcomes = ("CORRECT","CURATED","FILLED_IN", "UNABLE_DETERMINE_VALIDITY",  "UNABLE_CURATE") #col order in output
   #   print("outcomes=", outcomes)
###   validators = stats.getValidators()

#x   stats.stats2XLSX(workbook, worksheet, formats, validatorStats, origin1, outcomes, validators)
#   print("HELLO from L260\n")
#   stats.stats2XLSX(optdict)
   
   print("HELLO from L265\n")
  #### outcomestatsOnData(optdict)
#   workbook.close()
#   wb.close()
#   wb.save("styled2.xlsx")


def main():
   optdict = {}
   wb = Workbook()
   ws = wb.active #default worksheet in wb
   ocol = 4
   orow = 8
   thin = Side(border_style="thin", color="000000")
   double = Side(border_style="double", color="ff0000")
   border = Border(top=double, left=thin, right=thin, bottom=double)
   fill = PatternFill("solid", fgColor="DDDDDD")
   font = Font(b=True, color="FF0000")
   al = Alignment(horizontal="center", vertical="center")


   optdict['inputfile'] = './data/occurrence_qc.json'
   optdict['outputfile'] = 'outcomeStats.xlsx'
   optdict['workspace'] = './'
   optdict['configfile'] = './config/stats.ini'
#   optdict['loglevel'] = 'DEBUG'
   optdict['workbook'] = wb
   optdict['worksheet'] = ws
   optdict['origincolumn'] = ocol
   optdict['originrow'] = orow
   optdict['border'] = border
   optdict['fill'] = fill
   optdict['font'] = font
   optdict['alignment'] = al
   
#   print ('OutcomeStats L298 optdict: %s' % optdict)
###   stats = OutcomeStats(optdict)
   # formulate style
#   formatdict = {}
 #  frm = OutcomeFormats(formatdict)
###   frm = OutcomeFormats(optdict)
#   response=outcomestatsOnData(formatdict)
   response=outcomestatsOnData(optdict)
   print ('\nOutcomeStats response L300: %s' % response)
###   formats = frm.initFormats(frm)
   ws = wb.create_sheet("stats",0)  ##should be different openpyxl worksheet for each workflow?
   stats = OutcomeStats(optdict)
   stats.stats2XLSX(optdict)
   wb.save(optdict['outputfile'])

if __name__ == '__main__':
   main()
