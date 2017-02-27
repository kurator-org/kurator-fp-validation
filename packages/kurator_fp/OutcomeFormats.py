# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

__author__ = "Robert A. Morris"
__copyright__ = "Copyright 2016 President and Fellows of Harvard College"
__version__ = "OutcomeFormats.py 2017-02-26T18:43:14-0500"

import json
import sys
from openpyxl.styles import NamedStyle, PatternFill, Fill, Border, Side, Alignment, Protection, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
#from OpenpyxlStyle import style_range
from FormatUtils import style_range
import argparse

class OutcomeFormats:
   """Class supporting xlsx cell formats for a set of Kurator Quality Control *outcomes*
   """
   def __init__(self, outcomes):
      self.outcomes = outcomes
      
   def getFormats(self):
      return formats

   def setFormats(self, formats):
      return {}
   
#   def initFormats(self, workbook, worksheet):
   def initFormats(self, formatdict):  #ignore forrmatdict for now
      ### should get names of formats from something reachable from frmtdict
      ### perhaps via config/stats.ini
      formatGrnFill=PatternFill("solid", fgColor='00FF00') #lite green
      formatRedFill=PatternFill("solid", fgColor='FF0000')
      formatMusFill=PatternFill("solid", fgColor='DDDD00') #mustard
      formatYelFill=PatternFill("solid", fgColor='FFFF00')
      formatGryFill=PatternFill("solid", fgColor='888888')


      formatXFill=''
      #next make a dict out of the set of openpyxl styles
      self.formats={'UNABLE_DETERMINE_VALIDITY':formatGryFill, 'CURATED':formatYelFill, 'UNABLE_CURATE':formatRedFill, 'CORRECT':formatGrnFill, 'FILLED_IN':formatMusFill}

      

      return self.formats
def main():
   print("OutcomeFormats.main()")
   formatdict = {}
   frm = OutcomeFormats(formatdict)
   formats = frm.initFormats(formatdict)
   formatGrnFill=PatternFill("solid", fgColor='00FF00') #lite green
   formatRedFill=PatternFill("solid", fgColor='FF0000')
   formatMusFill=PatternFill("solid", fgColor='DDDD00') #mustard
   formatYelFill=PatternFill("solid", fgColor='222200')
   formatGryFill=PatternFill("solid", fgColor='888888')
   formatsDictX={'UNABLE\nDETERMINE\nVALIDITY':formatGryFill, 'CURATED':formatYelFill, 'UNABLE CURATE':formatRedFill, 'CORRECT':formatGrnFill, 'FILLED IN':formatMusFill}

   formatsDict={'CORRECT':formatGrnFill, 'CURATED':formatYelFill,'FILLED IN':formatMusFill, 'UNABLE CURATE':formatRedFill,' UNABLE\nDETERMINE\nVALIDITY':formatGryFill}
  
   wb = Workbook()
   ws = wb.active
   originrow = 4
   rd = ws.row_dimensions[originrow]
   rd.height = 12 #points; could do better if can compute out
   origincol = 2
   #headerColAl = Alignment()
   #headerColAl.vert = Alignment.VERT_CENTER

#   outcomes = eval(config['DEFAULT']['outcomes'])
#   max1= max(len(s) for s in validators)
#   max2= max(len(t) for t in outcomes)
#   maxlength = max(max1,max2)

   fdictlen = len(formatsDict)
#   print(fdictlen)
   theOutcomes = formatsDict.keys()
   max2 = len(max(formatsDict))
   max1=max2  #for now
#   print(max2, max3)
   

   
   outcomeIndex = 0

   # set outcome names as headers 
   #

#   font = Font(bold=True)
   for col in range(1, 1+fdictlen):
      value = theOutcomes[outcomeIndex]
      cell = ws.cell(column=col+origincol, row=originrow, value=value)
      thecol=col+origincol
      thecolletter = get_column_letter(thecol)
      print("L139 value = ", value, "thecolletter=",thecolletter, "col=", col, "cell=", cell)
#      cell.style.alignment.wrap_text = True
      cell.font = Font(bold=True)
      outcomeIndex += 1

      #set column width based on length of column header text
      #which is taken from outcome names
   dims = {}
   emwidth = 12 #for now
##   for row in ws.rows:
##      for cell in row:
##        if cell.value:
##            dims[cell.column] = emwidth + max((dims.get(cell.column, 0), len(cell.value)))
        #    print ("L92", cell.value, dims[cell.column])
#   maxcolwidth = max(dims.keys())
   ww = dims
   maxwidth = 0
   print("L149=",ww)
#   alignment=Alignment(horizontal='general',
#                       vertical='top',
#                       text_rotation=0,
#                       wrap_text=True,
#                       shrink_to_fit=False,
#                       indent=0)
   for col, value in dims.items():
      ws.column_dimensions[col].width = value
      if value > maxwidth :
         maxwidth = value
      print("L150 maxwidth=",maxwidth)
   for row in ws.rows:
#      for cell in row:
      for outcomename in row:
        if outcomename.value:
            dims[outcomename.column] = emwidth + maxwidth
#            cell.style.alignment.wrap_text = True
            
#            print("L161 outcomename= ",outcomename.value)


   validators = ("ScientificNameValidator","DateValidator",  "GeoRefValidator","BasisOfRecordValidator") #row order in output. should get from args
   numvalidators = len(validators)
   row = 1+originrow
 #  print(validators[row])

#   d=ws.cell(row=row, col=0, value=validators[row])
   i = 0
   col = 0

   maxValidatorLen = len(max(validators))
   print(maxValidatorLen)
   for i in range(0,numvalidators):
#      print(validators[row])
      value = validators[i]
      ws.cell(row=i+originrow+1, column=origincol, value=value)
   
   validatorCol = get_column_letter(origincol)
   print(validatorCol)
   ws.column_dimensions[validatorCol].width = maxValidatorLen
     
      # make cell color extracted from outcome index

   thin = Side(border_style="thin", color="000000")
   border = Border(top=thin, left=thin, right=thin, bottom=thin)
  # theFills = [PatternFill("solid", fgColor="00FF00"), PatternFill("solid", fgColor="FF0000"), PatternFill("solid", fgColor="DDDD00"), PatternFill("solid", fgColor="FFFF00"), PatternFill("solid", fgColor="BBBBBB")]
   theFills = [PatternFill("solid", fgColor="00FF00"), PatternFill("solid", fgColor="FFFF00"), PatternFill("solid", fgColor="DDDD00"), PatternFill("solid", fgColor="FF0000"), PatternFill("solid", fgColor="BBBBBB")]
   font = Font(b=True, color="000000")
   al = Alignment(horizontal="center", vertical="top", wrap_text=True)

   
   j=0
   for col in range(1+origincol,1+origincol+len(formatsDict)):
      colname = get_column_letter(col)
      yy = formatsDict.keys()[col-origincol-1]
#      zz = formatsDict.get(yy)
      print(col-origincol)
      theFill = theFills[j]
      print("L179 j=",j, "theFill=", theFill)
      j = j+1
      for row in range(1+originrow,1+originrow+numvalidators):
         cellname = colname+str(row)
         theRange=cellname+":"+cellname
         theCell = ws[cellname]
         style_range(ws,theRange,border=border, fill=theFill, font=font, alignment=al)

   

   wb.save("outcomestyled.xlsx")
   
if __name__ == "__main__" :
   main()

"""
      formatGrnFill=PatternFill("solid", fgColor='00FF00') #lite green
      formatRedFill=PatternFill("solid", fgColor='FF0000')
      formatMusFill=PatternFill("solid", fgColor='DDDD00') #mustard
      formatYelFill=PatternFill("solid", fgColor='FFFF00')
      formatGryFill=PatternFill("solid", fgColor='888888') 
      formatXFill=''
      #next make a dict out of the set of openpyxl styles
      self.formats={'UNABLE_DETERMINE_VALIDITY':formatGryFill, 'CURATED':formatYelFill, 'UNABLE_CURATE':formatRedFill, 'CORRECT':formatGrnFill, 'FILLED_IN':formatMusFill}
"""
