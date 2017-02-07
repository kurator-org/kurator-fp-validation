# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

__author__ = "Robert A. Morris"
__copyright__ = "Copyright 2016 President and Fellows of Harvard College"
__version__ = "OpenyxlStyle.py 2017-02-06T19:10:13-05:00"

import json
import sys
#import xlsxwriter
from openpyxl.styles import PatternFill, Fill, Border, Side, Alignment, Protection, Font, GradientFill, Alignment
from openpyxl import Workbook
import argparse

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Following http://openpyxl.readthedocs.io/en/default/styles.html#applying-styles
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

def main():
    print("In OpenpyxlStyle.main")
    wb = Workbook()
    ws = wb.active
    my_cell = ws['B2']
    my_cell.value = "My Cell"
    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="ff0000")

    border = Border(top=double, left=thin, right=thin, bottom=double)
    fill = PatternFill("solid", fgColor="DDDDDD")
    fill = GradientFill(stop=("000000", "FFFFFF"))
    font = Font(b=True, color="FF0000")
    al = Alignment(horizontal="center", vertical="center")

    style_range(ws, 'B2:F4', border=border, fill=fill, font=font, alignment=al)
    wb.save("styled.xlsx")

    print("Open styled.xlsx with a spreadsheet reader")
if __name__ == "__main__" :
   main()
