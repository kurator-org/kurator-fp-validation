# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

__author__ = "Robert A. Morris"
__copyright__ = "Copyright 2016 President and Fellows of Harvard College"
__version__ = "PythonToOpenPyxl.py 2017-05-24T11:09:22-04:00"

import json
import Config
import configparser
import argparse
import numpy as np
from openpyxl import Workbook
#from openpyxl.styles import.colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import OutcomeStats as ocstats
from OptDict import OptDict
import sys

def pythonToOpenpyxl(optdict):
    ''' This function is the actor 'onData' function invoked by the workflow engine. Reads
        occurrence data from a json inputfile and produces a stats report of outcomes for
        each validator.

    options - a dictionary of parameters
        loglevel - level at which to log (e.g., DEBUG) (optional)
        workspace - path to a directory for the outputfile (optional)
        inputfile - full path to the input file (required)
        outputfile - name of the output file, without path (optional)

        colOrigin - the column index at the origin of the stats table
        rowOrigin - the row index at the origin of the stats table

        outcomes - list of the names of outcomes (e.g. CORRECT, CURATED, etc)
        outcomeFills - dict that maps outcome names to corresponding fill color
        outcomesFolded - folded outcome names used as the column labels

        validators - list of validator names (e.g. ScientificNameValidator, etc)

    returns a dictionary with information about the results
        workspace - actual path to the directory where the outputfile was written
        outputfile - actual full path to the output file
        success - True if process completed successfully, otherwise False
        message - an explanation of the results
        artifacts - a dictionary of persistent objects created
    '''

    workspace = optdict['workspace']
    inputfile = optdict['inputfile']
    outputfile = optdict['outputfile']
    configfile = optdict['configfile']

    rowOrigin = optdict['rowOrigin']
    colOrigin = optdict['colOrigin']

    optdict = OptDict(workspace, inputfile, outputfile)  # init OptDict object
    optdict.configure(configfile)
    optdict.origin(rowOrigin, colOrigin)  # set origin to row 7, col 1

    # create the outcome stats workbook and sheet
    wb = Workbook()
    ws = wb.active

    setColumnStyles(ws, optdict)  # note: calls OutcomeStats.getStats(...) as well

    # save the workbook to the workspace directory
    outputfile = '%s/%s' % (workspace.rstrip('/'), outputfile)
    wb.save(outputfile)

    # TODO: report error conditions, for now the following will assume success
    success = True
    message = 'Successfully generated outcome stats report from occurrence json'

    # create a new OptDict instance for the response
    response = OptDict(workspace, inputfile, outputfile)
    # publish the artifact so that the workflow engine is aware of it
    response.publish('stats_report', outputfile)
    # report status of success or failure with descriptive text
    response.status(success, message)

    return response.optdict  # input dict with additions becomes the output dict

def wbInit():  #should call only once?
    wb = Workbook()
    return  wb

def getWorksheet(wb, sheet=None):
    if sheet is None:
        return wb.active
    else:
        return sheet

def vtl(string):
    return len(string)

def setColumnStyles(ws, optdict):
    thin   = Side(border_style="thin",   color="000000")
    double = Side(border_style="double", color="000000")
    border = Border(top=double,left=thin,right=thin,bottom=double)

    outcomes =   optdict['outcomes']
    validators = optdict['validators']
    outcomes2 =  optdict['outcomesFolded']
    colOrigin =  optdict['colOrigin']
    rowOrigin =  optdict['rowOrigin']
    outcomeFills = optdict['outcomeFills']
    maxvtl =  0 #max validator typographic length 
    for index in range(len(validators)): #enter validator labels and
                                         #find typographically longest
        value = validators[index]
        thecell=ws.cell(column=colOrigin, row=rowOrigin+index+1,
                        value=value)
           #next find the currently largest validator name string
        vtl = len(value)
        if vtl > maxvtl :
            maxvtl = vtl
        #reset column width to current value of maxvtl
    ws.column_dimensions[thecell.column].width = maxvtl


    maxotl = 0 #max outcome typographic length;
    for index in range(len(outcomes)):
        value = outcomes[index]
        otl = len(value)
        if otl>maxotl:
            maxotl = otl
    for index in range(len(outcomes)):
        value = outcomes2[index]
        otl = len(value) #GAAK! this is character count!
        alignment=Alignment(horizontal='general',
                     vertical='justify',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)
        thecell = ws.cell(column=colOrigin+index+1,row=rowOrigin,value=value)
        thecell.alignment = alignment
        thecell_col = thecell.column
        
        ws.column_dimensions[thecell_col].width = maxotl

    rd = ws.row_dimensions[rowOrigin]
    rd.height = 45
    statsAsTuples = ocstats.getStats(optdict) #from OpenStats module
    for index in range(len(statsAsTuples)):
        rownum=index
        row = rownum+1+rowOrigin
        theRowTuple = statsAsTuples[index]
        for index in range(len(theRowTuple)):
            colnum = index
            column = colnum+1+colOrigin
            value = theRowTuple[index]
            thecell = ws.cell(column=column,row=row,value=value)
            outcome = outcomes[colnum]
            thecell.fill =   PatternFill("solid", fgColor=outcomeFills[outcome])
            thecell.border = border


def main():
    optdict = OptDict('./', 'occurrence_qc.json', 'stats.xlsx')  # init OptDict object

    optdict.configure()  # default configuration is taken from stats.ini
    optdict.origin(7, 1)  # set origin to row 7, col 1

    # print input dict and call actor function
    print 'Input dictionary supplied as arg to pythonToOpenpyxl()'
    print optdict

    response = pythonToOpenpyxl(optdict)

    # print output dict from the response
    print 'Output dictionary returned from call to pythonToOpenpyxl()'
    print response

if __name__ == "__main__" :
   main()

